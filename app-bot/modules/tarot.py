# tarot.py
# tarot card delivery

import random
from datetime import datetime
import time

import psycopg2

import discord
from discord.ext import commands
from discord import app_commands

import re

import random

from cns import *

import rich

from rich import print as print

from openpyxl import Workbook
from openpyxl import load_workbook


class Tarot(commands.Cog):
    """
    Pull Tarot Cards! Try it in your DM with KREBBOT!
    tarot1 to pull one card, tarot3 to pull 3 cards!
    """

    def __init__(self, bot):
        self.bot = bot
        self._last_member = None
        self.intents = bot.intents
        # self.ui = UI(bot)
        self.filename = "database/exceldb/tarotraw.xlsx"

        #poll internet for images
        try:
            self.thumburlm = "https://p7.hiclipart.com/preview/189/56/52/sunlight-drawing-clip-art-sun-thumbnail.jpg"
        except:
            self.thumburlm = ""
        try:
            self.thumburls = "https://w7.pngwing.com/pngs/415/908/png-transparent-two-brown-and0-gray-swords-icon-sword-weapon-icon-swords-game-angle-samurai-sword-thumbnail.png"
        except:
            self.thumburls = ""
        try:
            self.thumburlw = "https://w7.pngwing.com/pngs/840/152/png-transparent-wand-magician-magic-wand-angle-leaf-symmetry-thumbnail.png"
        except:
            self.thumburlw = ""
        try:
            self.thumburlp = "https://png2.cleanpng.com/sh/92a688afad1cc1a1c575e569a6f6a477/L0KzQYm3VMIxN5xrfZH0aYP2gLBuTgBmdqVmfARqbT32ib7pjBwudZJsgdU2cHXxhLLukvFuNWZmfNMBY0PlR4XsUBIzNmI9SqMBMUC5QYa5VMI3O2oATaYCOEixgLBu/kisspng-pentagram-symbol-magic-pentagram-5ada6c3b74e3b2.1821610615242639954788.png"
        except:
            self.thumburlp = ""
        try:
            self.thumburlc = "https://e7.pngegg.com/pngimages/561/459/png-clipart-chalice-wine-glass-eucharist-paten-ciborium-cup-glass-wine-glass-thumbnail.png"
        except:
            self.thumburlc = ""

        self.tarotdata = []
        workbook = load_workbook(filename=self.filename)
        spreadsheet = workbook.active
        for i in range(78):
            row = i + 2
            data = []
            ID = spreadsheet.cell(row=row, column=1).value
            ARCANA_TYPE = spreadsheet.cell(row=row, column=2).value
            CARD_NAME = spreadsheet.cell(row=row, column=3).value
            ART_DESC = spreadsheet.cell(row=row, column=4).value
            UP_SUM = spreadsheet.cell(row=row, column=5).value
            UP_DESC = spreadsheet.cell(row=row, column=6).value
            REV_SUM = spreadsheet.cell(row=row, column=7).value
            REV_DESC = spreadsheet.cell(row=row, column=8).value
            PLANET = spreadsheet.cell(row=row, column=9).value
            SIGN = spreadsheet.cell(row=row, column=10).value
            ELEMENT = spreadsheet.cell(row=row, column=11).value

            data.append(ID)
            data.append(ARCANA_TYPE)
            data.append(CARD_NAME)
            data.append(ART_DESC)
            data.append(UP_SUM)
            data.append(UP_DESC)
            data.append(REV_SUM)
            data.append(REV_DESC)
            data.append(PLANET)
            data.append(SIGN)
            data.append(ELEMENT)

            self.tarotdata.append(data)

        print("Tarot Data Loaded.")

    async def buildcard(self, ctx, id: int, facing: str):
        facing = facing
        id = int(id)
        thumburl = ""

        if id <= 21:  # Major
            ccolor = discord.Color.blue()
            thumburl = self.thumburlm
        elif id > 21 and id <= 35:  # Swords
            ccolor = discord.Color.red()
            thumburl = self.thumburls
        elif id > 35 and id <= 49:  # Wands
            ccolor = discord.Color.orange()
            thumburl = self.thumburlw
        elif id > 50 and id <= 63:  # Pentacles
            ccolor = discord.Color.green()
            thumburl = self.thumburlp
        elif id > 63 and id <= 77:  # Cups
            ccolor = discord.Color.yellow()
            thumburl = self.thumburlc
        else:  # error
            ccolor = discord.Color.purple()
            thumburl = ""

        cd = self.tarotdata[id]
        embed = discord.Embed(title="KREBBOTAROT:", description="Enjoy your Tarot Card!", color=ccolor)
        embed.add_field(name="Card Name:", value=cd[2], inline=True)
        embed.add_field(name="Major/Minor Arcana:", value=cd[1], inline=True)
        if facing == "up":
            embed.add_field(name="Upright Summary", value=cd[4], inline=False)
            embed.add_field(name="Upright Details", value=cd[5], inline=False)
        if facing == "down":
            embed.add_field(name="Reverse Summary", value=cd[6], inline=False)
            embed.add_field(name="Reverse Details", value=cd[7], inline=False)
        embed.add_field(name="Associated Planet", value=cd[8], inline=True)
        embed.add_field(name="Associated Sign", value=cd[9], inline=True)
        embed.add_field(name="Associated Element", value=cd[10], inline=True)
        embed.set_thumbnail(url=thumburl)
        embed.set_footer(text="Delivered by KREBBOT with love <3")
        return embed
    

    @app_commands.command(name="tarot")
    @app_commands.guilds(discord.Object(id=754510720590151751))
    async def tarot(self, interaction: discord.Interaction, number_of_cards: int) -> None:
        """ Draws Tarot Cards! Enter the amount you want to draw in the command. """
        responseuser = interaction.user
        if number_of_cards > 10:
            number_of_cards = 10
        cards = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10,
                 11, 12, 13, 14, 16, 17, 18, 19, 20,
                    21, 22, 23, 24, 25, 26, 27, 28, 29,
                    30, 31, 32, 33, 34, 36, 37, 38, 39,
                    40, 41, 42, 43, 44, 45, 46, 47, 48,
                    49, 50, 51, 52, 53, 54, 55, 56, 57,
                    58, 59, 60, 61, 62, 63, 64, 65, 66,
                    67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77]
        cardbucket = []
        for i in range(number_of_cards):
            rcard = random.choice(cards)
            cards.remove(rcard)
            facechoice = ("up", "down")
            rfacing = random.choice(facechoice)
            card = await self.buildcard(interaction, id=rcard, facing=rfacing)
            response = "Okay " + responseuser.display_name + ", here is your card:"
            cardbucket.append(card)
            
        await interaction.response.send_message(response, embeds=cardbucket)

    @app_commands.command(name="fetchtarot")
    @app_commands.guilds(discord.Object(id=754510720590151751))
    async def fetchtarot(self, ctx, id: int, facing: str) -> None:
        """ Looks up a single card by ID and facing. For power users. """
        if await self.bot.is_prod_guild(ctx):
            return
        responseuser = ctx.message.author
        if responseuser.id != 181157857478049792:
            return
        rcard = id
        rfacing = facing
        card = await self.buildcard(ctx, id=rcard, facing=rfacing)
        response = "Okay " + responseuser.display_name + ", here is your card:"
        await ctx.channel.send(response)
        await ctx.channel.send(embed=card)
        pass

async def setup(bot):
    await bot.add_cog(Tarot(bot))