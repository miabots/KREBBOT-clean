# tarot.py
# tarot card delivery

import random
from datetime import datetime

import psycopg2

import discord
from discord.ext import commands

import re

import random

from cns import *

import rich

from rich import print as print

from openpyxl import Workbook
from openpyxl import load_workbook


class Tarot(commands.Cog):
    """
    Pull Tarot Cards! Try it in your DM with KREBBOT!
    tarot1 to pull one card, tarot3 to pull 3 cards!
    """

    def __init__(self, bot):
        self.bot = bot
        self._last_member = None
        self.intents = bot.intents
        #self.ui = UI(bot)
        self.filename = "database\exceldb\\tarotraw.xlsx"

    def readtarotexcel(self, ctx, row):
        workbook = load_workbook(filename=self.filename)
        sheetname = workbook.sheetnames
        spreadsheet = workbook.active

        data = []
        ID = spreadsheet.cell(row=row, column=1).value
        ARCANA_TYPE = spreadsheet.cell(row=row, column=2).value
        CARD_NAME = spreadsheet.cell(row=row, column=3).value
        ART_DESC = spreadsheet.cell(row=row, column=4).value
        UP_SUM = spreadsheet.cell(row=row, column=5).value
        UP_DESC = spreadsheet.cell(row=row, column=6).value
        REV_SUM = spreadsheet.cell(row=row, column=7).value
        REV_DESC = spreadsheet.cell(row=row, column=8).value
        PLANET = spreadsheet.cell(row=row, column=9).value
        SIGN = spreadsheet.cell(row=row, column=10).value
        ELEMENT = spreadsheet.cell(row=row, column=11).value

        data.append(ID)
        data.append(ARCANA_TYPE)
        data.append(CARD_NAME)
        data.append(ART_DESC)
        data.append(UP_SUM)
        data.append(UP_DESC)
        data.append(REV_SUM)
        data.append(REV_DESC)
        data.append(PLANET)
        data.append(SIGN)
        data.append(ELEMENT)

        length = len(data)

        # for item in data:
        #    print(item)

        # for i in range(length):
        #    print(data[i])

        return data
        pass

        # tarot IDs are 0 to 77

    async def buildcard(self, ctx, id: int, facing: str):

        facing = facing

        id = int(id)

        thumburl = ""

        if (id <= 21):  # Major
            ccolor = discord.Color.blue()
            try:
                thumburl = "https://p7.hiclipart.com/preview/189/56/52/sunlight-drawing-clip-art-sun-thumbnail.jpg"
            except:
                thumburl = ""
            # print("1")
        elif (id > 21 and id <= 35):  # Swords
            ccolor = discord.Color.red()
            try:
                thumburl = "https://w7.pngwing.com/pngs/415/908/png-transparent-two-brown-and0-gray-swords-icon-sword-weapon-icon-swords-game-angle-samurai-sword-thumbnail.png"
            except:
                thumburl = ""

            # print("2")
        elif (id > 35 and id <= 49):  # Wands
            ccolor = discord.Color.orange()
            try:
                thumburl = "https://w7.pngwing.com/pngs/840/152/png-transparent-wand-magician-magic-wand-angle-leaf-symmetry-thumbnail.png"
            except:
                thumburl = ""
            # print("3")
        elif (id > 50 and id <= 63):  # Pentacles
            ccolor = discord.Color.green()
            try:
                thumburl = "https://upload.wikimedia.org/wikipedia/commons/thumb/a/ac/Pentacle_on_white.svg/480px-Pentacle_on_white.svg.png"
            except:
                thumburl = ""

            # print("4")
        elif (id > 63 and id <= 77):  # Cups
            ccolor = discord.Color.yellow()
            try:
                thumburl = "https://e7.pngegg.com/pngimages/561/459/png-clipart-chalice-wine-glass-eucharist-paten-ciborium-cup-glass-wine-glass-thumbnail.png"
            except:
                thumburl = ""
            # print("5")
        else:  # error
            ccolor = discord.Color.purple()
            try:
                thumburl = ""
            except:
                thumburl = ""
            # print("6")

        #print("ID AND COLOR")
        # print(id)
        # print(ccolor)

        row = id+2
        cd = self.readtarotexcel(ctx=ctx, row=row)
        embed = discord.Embed(title="KREBBOTAROT:", description="Enjoy your Tarot Card!", color=ccolor)
        # embed.set_author(name="TEST AUTHOR")
        embed.add_field(name="Card Name:", value=cd[2], inline=True)
        embed.add_field(name="Major/Minor Arcana:", value=cd[1], inline=True)
        if(facing == "up"):
            embed.add_field(name="Upright Summary", value=cd[4], inline=False)
            embed.add_field(name="Upright Details", value=cd[5], inline=False)
        if(facing == "down"):
            embed.add_field(name="Reverse Summary", value=cd[6], inline=False)
            embed.add_field(name="Reverse Details", value=cd[7], inline=False)
        embed.add_field(name="Associated Planet", value=cd[8], inline=True)
        embed.add_field(name="Associated Sign", value=cd[9], inline=True)
        embed.add_field(name="Associated Element", value=cd[10], inline=True)
        # embed.set_image(url=url)
        embed.set_thumbnail(url=thumburl)
        embed.set_footer(text="Delivered by KREBBOT with love <3")
        # embed = discord.Embed(title=title, description=desc)
        # embed.set_author(name="TEST AUTHOR")
        # embed.add_field(name="TestFieldName", value="TestFieldValue", inline=False)
        # embed.set_image(url=url)
        # embed.set_footer(text="Delivered by KREBBOT with love <3")

        # await ctx.channel.send(embed=embed)

        return embed

    @ commands.command(help="Draws One Random Tarot Card", brief="Tarot")
    async def tarot1(self, ctx):
        if await self.bot.is_prod_guild(ctx):
            return
        responseuser = ctx.message.author
        capctx = ctx.message.channel
        rcard = random.randint(0, 77)
        # print("RCARD")
        # print(rcard)
        facechoice = ("up", "down")
        rfacing = random.choice(facechoice)
        card = await self.buildcard(ctx, id=rcard, facing=rfacing)
        # print("CARD")
        # print(card)
        response = "Okay " + responseuser.display_name + ", here is your card:"
        await ctx.channel.send(response)
        await ctx.channel.send(embed=card)
        pass

    @ commands.command(help="Draws Three Random Tarot Cards at once", brief="Tarot")
    async def tarot3(self, ctx):
        if await self.bot.is_prod_guild(ctx):
            return
        responseuser = ctx.message.author
        capctx = ctx.message.channel
        rcard = random.randint(0, 77)
        facechoice = ("up", "down")
        rfacing = random.choice(facechoice)
        card = await self.buildcard(ctx, id=rcard, facing=rfacing)
        rcard = random.randint(0, 77)
        facechoice = ("up", "down")
        rfacing = random.choice(facechoice)
        card2 = await self.buildcard(ctx, id=rcard, facing=rfacing)
        rcard = random.randint(0, 77)
        facechoice = ("up", "down")
        rfacing = random.choice(facechoice)
        card3 = await self.buildcard(ctx, id=rcard, facing=rfacing)
        response = "Okay " + responseuser.display_name + ", here is your 3 card draw:"
        await ctx.channel.send(response)
        await ctx.channel.send(embed=card)
        await ctx.channel.send(embed=card2)
        await ctx.channel.send(embed=card3)
        pass


async def setup(bot):
    await bot.add_cog(Tarot(bot))
